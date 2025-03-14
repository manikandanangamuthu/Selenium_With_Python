import openpyxl

from openpyxl.styles import PatternFill

def getRowCount(file_path,shaeetName):
    workbook=openpyxl.load_workbook(file_path)
    sheet=workbook[shaeetName]
    return (sheet.max_row)
def getColumnCount(file_Path,sheetName):
    workbook=openpyxl.load_workbook(file_Path)
    sheet=workbook[sheetName]
    return (sheet.max_column)
def readDate(file_Path,sheetName,rownum,columnnum):
    workbook=openpyxl.load_workbook(file_Path)
    sheet=workbook[sheetName]
    return sheet.cell(rownum,columnnum).value

def writeData(file_Path,sheetName,rownum,columnnum):
    workbook=openpyxl.load_workbook(file_Path)
    sheet=workbook[sheetName]
    sheet.cell(rownum,columnnum).value = data
    workbook.save(file_Path)

def fillGreenColor(file_Path,sheetName,rownum,columnnum):
    workbook=openpyxl.load_workbook(file_Path)

    sheet=workbook[sheetName]
    greenFill= patternFill(start_color='60b212',
                           end_color='60b212',
                           fill_type='solid')
    sheet.cell(rownum,columnnum).fill=greenFill
    workbook.save(file_Path)


def fillRedColor(file_Path, sheetName, rownum, columnnum):
    workbook = openpyxl.load_workbook(file_Path)

    sheet = workbook[sheetName]
    redFill = patternFill(start_color='ff0000',
                            end_color='ff0000',
                            fill_type='solid')
    sheet.cell(rownum, columnnum).fill = redFill
    workbook.save(file_Path)







