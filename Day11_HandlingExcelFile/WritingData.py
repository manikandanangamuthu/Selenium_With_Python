import openpyxl

# # Same Data
# file_path="C:\\Users\\05444\\Desktop\\Data.xlsx"
# workbook= openpyxl.load_workbook(file_path)
#
# sheet=workbook.active #or sheet=workbook["Sheet1"]   -->  active sheet will use for only one sheet if we have in excel fle
# for r in range(1,6):
#     for c in range(1,4):
#         sheet.cell(r, c).value="welcome"
# workbook.save(file_path)


# Multiple data

file_path="C:\\Users\\05444\\Desktop\\Data.xlsx"
workbook= openpyxl.load_workbook(file_path)

# sheet=workbook("Sheet1")
sheet=workbook.active

sheet.cell(1,1).value="S.No"
sheet.cell(1,2).value="Name"
sheet.cell(1,3).value="Emp.ID"

sheet.cell(2,1).value=1
sheet.cell(2,2).value="mani"
sheet.cell(2,3).value=1

sheet.cell(3,1).value=2
sheet.cell(3,2).value="khem"
sheet.cell(3,3).value=2

sheet.cell(4,1).value=3
sheet.cell(4,2).value="niraj"
sheet.cell(4,3).value=3
workbook.save(file_path)  # save the file after entring the data