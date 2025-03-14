import openpyxl
# File -->> Workbook --> sheets --> rows -->> cells(column)
file_path="C:\\Users\\05444\\Desktop\\Exam_Slot1.xlsx"

workbook=openpyxl.load_workbook(file_path)

sheet=workbook["Sheet4"]

rows=sheet.max_row # count number of rows in the excel sheet
cols=sheet.max_column # count number of column in the excel sheet

# reading all the column & rows from the excel sheet

for r in range(1,rows+1):
    for c in range(1,cols+1):
        print(sheet.cell(r,c).value,end='      ')
    print()
